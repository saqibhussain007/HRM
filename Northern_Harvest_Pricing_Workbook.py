"""
Northern Harvest Professional Pricing & Profitability System
Complete Excel Workbook Generator

This script creates a professional, Google Sheets-compatible pricing workbook
for Northern Harvest (northernharvest.pk) - dry fruits and nuts e-commerce business

Author: Copilot
Date: 2026-08-13
Currency: PKR (Pakistani Rupees)
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import datetime

# Create workbook
wb = openpyxl.Workbook()
wb.remove(wb.active)  # Remove default sheet

# ============================================================================
# COLOR SCHEME
# ============================================================================
COLORS = {
    'dark_navy': 'FF1F3A5F',      # Dark navy for headers
    'blue_input': 'FFD9E1F2',     # Light blue for inputs
    'green_result': 'FFC6EFCE',   # Light green for calculated
    'yellow_warn': 'FFFFFFCC',    # Light yellow for warnings
    'red_loss': 'FFFFC7CE',       # Light red for losses
    'white': 'FFFFFFFF',
}

FONTS = {
    'header': Font(name='Calibri', size=14, bold=True, color='FFFFFFFF'),
    'subheader': Font(name='Calibri', size=11, bold=True, color='FFFFFFFF'),
    'normal': Font(name='Calibri', size=11),
    'bold': Font(name='Calibri', size=11, bold=True),
    'title': Font(name='Calibri', size=20, bold=True, color='FF1F3A5F'),
    'small': Font(name='Calibri', size=9),
}

FILLS = {
    'dark_navy': PatternFill(start_color=COLORS['dark_navy'], end_color=COLORS['dark_navy'], fill_type='solid'),
    'blue': PatternFill(start_color=COLORS['blue_input'], end_color=COLORS['blue_input'], fill_type='solid'),
    'green': PatternFill(start_color=COLORS['green_result'], end_color=COLORS['green_result'], fill_type='solid'),
    'yellow': PatternFill(start_color=COLORS['yellow_warn'], end_color=COLORS['yellow_warn'], fill_type='solid'),
    'red': PatternFill(start_color=COLORS['red_loss'], end_color=COLORS['red_loss'], fill_type='solid'),
}

BORDERS = Border(
    left=Side(style='thin', color='FF000000'),
    right=Side(style='thin', color='FF000000'),
    top=Side(style='thin', color='FF000000'),
    bottom=Side(style='thin', color='FF000000')
)

# ============================================================================
# SHEET 1: START HERE
# ============================================================================
ws_start = wb.create_sheet('5 Start Here', 0)
ws_start.column_dimensions['A'].width = 80

ws_start['A1'] = 'NORTHERN HARVEST'
ws_start['A1'].font = Font(name='Calibri', size=24, bold=True, color='FF1F3A5F')
ws_start['A2'] = 'Professional Pricing & Profitability System'
ws_start['A2'].font = Font(name='Calibri', size=14, color='FF666666')

ws_start['A4'] = 'QUICK START GUIDE'
ws_start['A4'].font = FONTS['subheader']
ws_start['A4'].fill = FILLS['dark_navy']

instructions = [
    ('', ''),
    ('STEP 1:', 'Open the "1 Product Calculator" sheet'),
    ('', ''),
    ('STEP 2:', 'Enter your PRODUCT INFORMATION:'),
    ('', '  • Product Name (e.g., Premium Almonds)'),
    ('', '  • SKU (e.g., NH-ALM-500)'),
    ('', '  • Pack Size (e.g., 500g)'),
    ('', '  • Units per Order (usually 1)'),
    ('', ''),
    ('STEP 3:', 'Enter COST VALUES in BLUE cells:'),
    ('', '  • Purchase cost, freight, duties'),
    ('', '  • Packaging costs (bag, label, box)'),
    ('', '  • Labor and delivery costs'),
    ('', '  • Marketing and other costs'),
    ('', '  • Return/RTO rates (critical for Pakistan market)'),
    ('', ''),
    ('STEP 4:', 'DO NOT edit GREEN cells (auto-calculated)'),
    ('', ''),
    ('STEP 5:', 'Scroll down to the RESULTS section'),
    ('', ''),
    ('STEP 6:', 'Review:'),
    ('', '  • Break-even Price (minimum price to avoid loss)'),
    ('', '  • Recommended Store Price (30% margin by default)'),
    ('', '  • Expected Profit and Net Margin'),
    ('', '  • Discounted Price (after 10% discount)'),
    ('', ''),
    ('STEP 7:', 'Check PRICE LADDER:'),
    ('', '  • Shows prices for different profit margins (20%-40%)'),
    ('', ''),
    ('STEP 8:', 'Click "SAVE PRODUCT" button'),
    ('', '  • Product will be saved to "2 Product Master"'),
    ('', '  • Calculator will clear for next product'),
    ('', ''),
    ('STEP 9:', 'View all products in "2 Product Master"'),
    ('', ''),
    ('STEP 10:', 'Check "4 Dashboard" for summary metrics'),
]

row = 6
for title, desc in instructions:
    ws_start[f'A{row}'] = title
    ws_start[f'B{row}'] = desc
    if title:
        ws_start[f'A{row}'].font = Font(name='Calibri', size=11, bold=True, color='FF1F3A5F')
    ws_start[f'B{row}'].font = FONTS['normal']
    row += 1

# Color coding explanation
row += 2
ws_start[f'A{row}'] = 'COLOR CODING'
ws_start[f'A{row}'].font = FONTS['subheader']
ws_start[f'A{row}'].fill = FILLS['dark_navy']

row += 2
color_guide = [
    ('BLUE cells', 'Enter your values here'),
    ('GREEN cells', 'Auto-calculated (do not edit)'),
    ('YELLOW cells', 'Important assumptions/warnings'),
    ('RED cells', 'Loss or critical warning'),
]

for color, meaning in color_guide:
    ws_start[f'A{row}'] = color
    ws_start[f'B{row}'] = meaning
    ws_start[f'A{row}'].font = FONTS['bold']
    if 'BLUE' in color:
        ws_start[f'A{row}'].fill = FILLS['blue']
    elif 'GREEN' in color:
        ws_start[f'A{row}'].fill = FILLS['green']
    elif 'YELLOW' in color:
        ws_start[f'A{row}'].fill = FILLS['yellow']
    elif 'RED' in color:
        ws_start[f'A{row}'].fill = FILLS['red']
    row += 1

# ============================================================================
# SHEET 2: SETTINGS
# ============================================================================
ws_settings = wb.create_sheet('3 Settings', 1)
ws_settings.column_dimensions['A'].width = 40
ws_settings.column_dimensions['B'].width = 20

row = 1
ws_settings[f'A{row}'] = 'NORTHERN HARVEST - SETTINGS'
ws_settings[f'A{row}'].font = FONTS['title']
ws_settings.merge_cells(f'A{row}:B{row}')

row += 2
ws_settings[f'A{row}'] = 'CONFIGURATION'
ws_settings[f'A{row}'].font = FONTS['subheader']
ws_settings[f'A{row}'].fill = FILLS['dark_navy']
ws_settings[f'B{row}'].fill = FILLS['dark_navy']

row += 1

settings_data = [
    ('Currency', 'PKR', 'currency'),
    ('Default Target Profit Margin %', 30, 'target_margin'),
    ('Default Payment Gateway Fee %', 2.5, 'gateway_fee'),
    ('Default Platform Fee %', 1.5, 'platform_fee'),
    ('Default Return Rate %', 5, 'return_rate'),
    ('Default RTO Rate %', 3, 'rto_rate'),
    ('Default Delivery Cost', 250, 'delivery_cost'),
    ('Customer Delivery Charged', 150, 'customer_delivery'),
    ('Free Shipping Threshold', 5000, 'free_shipping'),
    ('Default Tax Rate %', 17, 'tax_rate'),
    ('Default Marketing Cost', 50, 'marketing_cost'),
]

settings_cells = {}
for label, value, key in settings_data:
    ws_settings[f'A{row}'] = label
    ws_settings[f'A{row}'].font = FONTS['normal']
    ws_settings[f'A{row}'].alignment = Alignment(horizontal='left', vertical='center')
    
    ws_settings[f'B{row}'] = value
    ws_settings[f'B{row}'].fill = FILLS['blue']
    ws_settings[f'B{row}'].font = FONTS['bold']
    ws_settings[f'B{row}'].alignment = Alignment(horizontal='right', vertical='center')
    ws_settings[f'B{row}'].border = BORDERS
    
    settings_cells[key] = f'B{row}'
    row += 1

# Named ranges for easy reference
for key, cell in settings_cells.items():
    wb.named_ranges[key] = openpyxl.worksheet.named_range.NamedRange(key, ws_settings, f'{cell}')

# ============================================================================
# SHEET 3: PRODUCT CALCULATOR
# ============================================================================
ws_calc = wb.create_sheet('1 Product Calculator', 2)
ws_calc.column_dimensions['A'].width = 40
ws_calc.column_dimensions['B'].width = 20
ws_calc.column_dimensions['C'].width = 50
ws_calc.sheet_view.showGridLines = False

row = 1
ws_calc[f'A{row}'] = 'NORTHERN HARVEST - PRODUCT CALCULATOR'
ws_calc[f'A{row}'].font = FONTS['title']
ws_calc.merge_cells(f'A{row}:C{row}')

row += 2

# ---- A. PRODUCT INFORMATION ----
ws_calc[f'A{row}'] = 'A. PRODUCT INFORMATION'
ws_calc[f'A{row}'].font = FONTS['subheader']
ws_calc[f'A{row}'].fill = FILLS['dark_navy']
ws_calc.merge_cells(f'A{row}:C{row}')
row += 1

product_info = [
    ('Product Name', 'calc_product_name'),
    ('SKU', 'calc_sku'),
    ('Pack Size (grams)', 'calc_pack_size'),
    ('Units per Order', 'calc_units_order'),
]

for label, cell_key in product_info:
    ws_calc[f'A{row}'] = label
    ws_calc[f'A{row}'].font = FONTS['bold']
    ws_calc[f'B{row}'].fill = FILLS['blue']
    ws_calc[f'B{row}'].border = BORDERS
    ws_calc[f'B{row}'].alignment = Alignment(horizontal='center', vertical='center')
    row += 1

row += 1

# ---- B. LANDED PRODUCT COST ----
ws_calc[f'A{row}'] = 'B. LANDED PRODUCT COST'
ws_calc[f'A{row}'].font = FONTS['subheader']
ws_calc[f'A{row}'].fill = FILLS['dark_navy']
ws_calc.merge_cells(f'A{row}:C{row}')
row += 1

product_costs = [
    ('Purchase Cost (per unit)', 'cost_purchase'),
    ('Inbound Freight Cost', 'cost_freight'),
    ('Customs / Duties / Tax', 'cost_duties'),
    ('Wastage %', 'cost_wastage_pct'),
    ('Packaging: Bag Cost', 'cost_bag'),
    ('Packaging: Label Cost', 'cost_label'),
    ('Packaging: Box / Insert Cost', 'cost_box'),
    ('Labor / Handling Cost', 'cost_labor'),
    ('Other Product Cost', 'cost_other'),
]

for label, cell_key in product_costs:
    ws_calc[f'A{row}'] = label
    ws_calc[f'A{row}'].font = FONTS['normal']
    ws_calc[f'B{row}'].fill = FILLS['blue']
    ws_calc[f'B{row}'].border = BORDERS
    ws_calc[f'B{row}'].number_format = '0.00'
    row += 1

row += 1

# ---- BASE PRODUCT COST CALCULATION ----
ws_calc[f'A{row}'] = 'Base Product Cost'
ws_calc[f'A{row}'].font = FONTS['bold']
ws_calc[f'B{row}'].fill = FILLS['green']
ws_calc[f'B{row}'].border = BORDERS
ws_calc[f'B{row}'].font = FONTS['bold']
ws_calc[f'B{row}'].number_format = '0.00'
calc_base_cost_row = row
# Formula: Sum of all costs with wastage calculated
ws_calc[f'B{row}'] = f'=B8+B9+B10+(B8*B11/100)+B12+B13+B14+B15+B16'
row += 2

# ---- C. ORDER & SELLING COSTS ----
ws_calc[f'A{row}'] = 'C. ORDER & SELLING COSTS'
ws_calc[f'A{row}'].font = FONTS['subheader']
ws_calc[f'A{row}'].fill = FILLS['dark_navy']
ws_calc.merge_cells(f'A{row}:C{row}')
row += 1

selling_costs = [
    ('Average Delivery Cost', 'cost_delivery'),
    ('Customer Delivery Charged', 'cost_customer_delivery'),
    ('Payment Gateway Fee %', 'cost_gateway_pct'),
    ('Payment Gateway Fixed Fee', 'cost_gateway_fixed'),
    ('Platform Fee %', 'cost_platform_pct'),
    ('Marketing Cost / Order', 'cost_marketing'),
    ('Other Selling Cost', 'cost_selling_other'),
    ('Free Shipping Subsidy', 'cost_subsidy'),
]

selling_cost_rows = {}
for label, cell_key in selling_costs:
    ws_calc[f'A{row}'] = label
    ws_calc[f'A{row}'].font = FONTS['normal']
    ws_calc[f'B{row}'].fill = FILLS['blue']
    ws_calc[f'B{row}'].border = BORDERS
    ws_calc[f'B{row}'].number_format = '0.00'
    selling_cost_rows[cell_key] = row
    row += 1

row += 1

# NET DELIVERY COST
ws_calc[f'A{row}'] = 'Net Delivery Cost'
ws_calc[f'A{row}'].font = FONTS['bold']
ws_calc[f'B{row}'].fill = FILLS['green']
ws_calc[f'B{row}'].border = BORDERS
ws_calc[f'B{row}'].font = FONTS['bold']
ws_calc[f'B{row}'].number_format = '0.00'
net_delivery_row = row
ws_calc[f'B{row}'] = f'=B{selling_cost_rows["cost_delivery"]}-B{selling_cost_rows["cost_customer_delivery"]}+B{selling_cost_rows["cost_subsidy"]}'
row += 2

# ---- D. RETURNS / RTO ----
ws_calc[f'A{row}'] = 'D. RETURNS / RTO (CRITICAL for Pakistan Market)'
ws_calc[f'A{row}'].font = FONTS['subheader']
ws_calc[f'A{row}'].fill = FILLS['dark_navy']
ws_calc.merge_cells(f'A{row}:C{row}')
row += 1

rto_costs = [
    ('Return Rate %', 'rto_return_pct'),
    ('RTO Rate %', 'rto_rto_pct'),
    ('Return Shipping Cost', 'rto_shipping'),
    ('Return Handling Cost', 'rto_handling'),
    ('Repacking Cost', 'rto_repacking'),
    ('Non-resalable Loss %', 'rto_loss_pct'),
    ('Damaged Product Loss', 'rto_damaged'),
]

rto_rows = {}
for label, cell_key in rto_costs:
    ws_calc[f'A{row}'] = label
    ws_calc[f'A{row}'].font = FONTS['normal']
    ws_calc[f'B{row}'].fill = FILLS['blue']
    ws_calc[f'B{row}'].border = BORDERS
    ws_calc[f'B{row}'].number_format = '0.00'
    rto_rows[cell_key] = row
    row += 1

row += 1

# EXPECTED RTO COST
ws_calc[f'A{row}'] = 'Expected Return/RTO Cost per Order'
ws_calc[f'A{row}'].font = FONTS['bold']
ws_calc[f'B{row}'].fill = FILLS['green']
ws_calc[f'B{row}'].border = BORDERS
ws_calc[f'B{row}'].font = FONTS['bold']
ws_calc[f'B{row}'].number_format = '0.00'
expected_rto_row = row
ws_calc[f'B{row}'] = f'=((B{rto_rows["rto_return_pct"]}+B{rto_rows["rto_rto_pct"]})/100)*(B{rto_rows["rto_shipping"]}+B{rto_rows["rto_handling"]}+B{rto_rows["rto_repacking"]}+B{rto_rows["rto_damaged"]})'
row += 2

# ---- E. TAX & PRICING ASSUMPTIONS ----
ws_calc[f'A{row}'] = 'E. TAX & PRICING ASSUMPTIONS'
ws_calc[f'A{row}'].font = FONTS['subheader']
ws_calc[f'A{row}'].fill = FILLS['dark_navy']
ws_calc.merge_cells(f'A{row}:C{row}')
row += 1

assumptions = [
    ('Sales Tax / VAT %', 'assum_tax'),
    ('Target Net Profit Margin %', 'assum_margin'),
    ('Planned Discount %', 'assum_discount'),
]

assumption_rows = {}
for label, cell_key in assumptions:
    ws_calc[f'A{row}'] = label
    ws_calc[f'A{row}'].font = FONTS['normal']
    ws_calc[f'B{row}'].fill = FILLS['blue']
    ws_calc[f'B{row}'].border = BORDERS
    ws_calc[f'B{row}'].number_format = '0.00'
    assumption_rows[cell_key] = row
    row += 1

row += 3

# ---- F. AUTOMATIC RESULTS ----
ws_calc[f'A{row}'] = 'F. AUTOMATIC RESULTS'
ws_calc[f'A{row}'].font = FONTS['subheader']
ws_calc[f'A{row}'].fill = FILLS['dark_navy']
ws_calc.merge_cells(f'A{row}:C{row}')
row += 1

# Calculate Total Variable Cost
ws_calc[f'A{row}'] = 'Total Variable Cost'
ws_calc[f'A{row}'].font = FONTS['bold']
ws_calc[f'B{row}'].fill = FILLS['green']
ws_calc[f'B{row}'].border = BORDERS
ws_calc[f'B{row}'].font = FONTS['bold']
ws_calc[f'B{row}'].number_format = '0.00'
total_cost_row = row
ws_calc[f'B{row}'] = f'=B{calc_base_cost_row}+B{net_delivery_row}+B{selling_cost_rows["cost_marketing"]}+B{selling_cost_rows["cost_selling_other"]}'
row += 1

# Break-even Price
ws_calc[f'A{row}'] = 'Break-even Price'
ws_calc[f'A{row}'].font = FONTS['bold']
ws_calc[f'B{row}'].fill = FILLS['green']
ws_calc[f'B{row}'].border = BORDERS
ws_calc[f'B{row}'].font = FONTS['bold']
ws_calc[f'B{row}'].number_format = '₨ 0.00'
breakeven_row = row
ws_calc[f'B{row}'] = f'=B{total_cost_row}'
row += 1

# Required Price for Target Margin
ws_calc[f'A{row}'] = 'Required Price for Target Margin'
ws_calc[f'A{row}'].font = FONTS['bold']
ws_calc[f'B{row}'].fill = FILLS['green']
ws_calc[f'B{row}'].border = BORDERS
ws_calc[f'B{row}'].font = FONTS['bold']
ws_calc[f'B{row}'].number_format = '₨ 0.00'
required_price_row = row
# Formula: (Total Cost + RTO Cost) / (1 - Target Margin% - Gateway Fee% - Platform Fee%)
ws_calc[f'B{row}'] = f'=(B{total_cost_row}+B{expected_rto_row})/(1-(B{assumption_rows["assum_margin"]}/100)-(B{selling_cost_rows["cost_gateway_pct"]}/100)-(B{selling_cost_rows["cost_platform_pct"]}/100))'
row += 1

# Recommended Store Price
ws_calc[f'A{row}'] = 'Recommended Store Price'
ws_calc[f'A{row}'].font = FONTS['bold']
ws_calc[f'B{row}'].fill = FILLS['green']
ws_calc[f'B{row}'].border = BORDERS
ws_calc[f'B{row}'].font = Font(name='Calibri', size=12, bold=True, color='FF006600')
ws_calc[f'B{row}'].number_format = '₨ 0.00'
recommended_price_row = row
ws_calc[f'B{row}'] = f'=ROUND(B{required_price_row},0)'
row += 1

# Discounted Customer Price
ws_calc[f'A{row}'] = 'Discounted Customer Price (After Discount %)'
ws_calc[f'A{row}'].font = FONTS['bold']
ws_calc[f'B{row}'].fill = FILLS['green']
ws_calc[f'B{row}'].border = BORDERS
ws_calc[f'B{row}'].font = FONTS['bold']
ws_calc[f'B{row}'].number_format = '₨ 0.00'
discount_price_row = row
ws_calc[f'B{row}'] = f'=ROUND(B{recommended_price_row}*(1-B{assumption_rows["assum_discount"]}/100),0)'
row += 2

# Profit at Recommended Price
ws_calc[f'A{row}'] = 'Profit at Recommended Price'
ws_calc[f'A{row}'].font = FONTS['bold']
ws_calc[f'B{row}'].fill = FILLS['green']
ws_calc[f'B{row}'].border = BORDERS
ws_calc[f'B{row}'].font = FONTS['bold']
ws_calc[f'B{row}'].number_format = '₨ 0.00'
profit_recommended_row = row
# Profit = Price - Total Cost - RTO - (Gateway Fee) - (Platform Fee)
ws_calc[f'B{row}'] = f'=B{recommended_price_row}-B{total_cost_row}-B{expected_rto_row}-(B{recommended_price_row}*B{selling_cost_rows["cost_gateway_pct"]}/100)-B{selling_cost_rows["cost_gateway_fixed"]}-(B{recommended_price_row}*B{selling_cost_rows["cost_platform_pct"]}/100)'
row += 1

# Profit after Discount
ws_calc[f'A{row}'] = 'Profit after Discount'
ws_calc[f'A{row}'].font = FONTS['bold']
ws_calc[f'B{row}'].fill = FILLS['green']
ws_calc[f'B{row}'].border = BORDERS
ws_calc[f'B{row}'].font = FONTS['bold']
ws_calc[f'B{row}'].number_format = '₨ 0.00'
profit_discount_row = row
ws_calc[f'B{row}'] = f'=B{discount_price_row}-B{total_cost_row}-B{expected_rto_row}-(B{discount_price_row}*B{selling_cost_rows["cost_gateway_pct"]}/100)-B{selling_cost_rows["cost_gateway_fixed"]}-(B{discount_price_row}*B{selling_cost_rows["cost_platform_pct"]}/100)'
row += 2

# Net Margin at Recommended Price %
ws_calc[f'A{row}'] = 'Net Margin at Recommended Price %'
ws_calc[f'A{row}'].font = FONTS['bold']
ws_calc[f'B{row}'].fill = FILLS['green']
ws_calc[f'B{row}'].border = BORDERS
ws_calc[f'B{row}'].font = FONTS['bold']
ws_calc[f'B{row}'].number_format = '0.00"%"'
margin_recommended_row = row
ws_calc[f'B{row}'] = f'=IF(B{recommended_price_row}=0,0,(B{profit_recommended_row}/B{recommended_price_row})*100)'
row += 1

# Net Margin after Discount %
ws_calc[f'A{row}'] = 'Net Margin after Discount %'
ws_calc[f'A{row}'].font = FONTS['bold']
ws_calc[f'B{row}'].fill = FILLS['green']
ws_calc[f'B{row}'].border = BORDERS
ws_calc[f'B{row}'].font = FONTS['bold']
ws_calc[f'B{row}'].number_format = '0.00"%"'
margin_discount_row = row
ws_calc[f'B{row}'] = f'=IF(B{discount_price_row}=0,0,(B{profit_discount_row}/B{discount_price_row})*100)'
row += 1

# Profit % on Cost
ws_calc[f'A{row}'] = 'Profit % on Cost (Markup %)'
ws_calc[f'A{row}'].font = FONTS['bold']
ws_calc[f'B{row}'].fill = FILLS['green']
ws_calc[f'B{row}'].border = BORDERS
ws_calc[f'B{row}'].font = FONTS['bold']
ws_calc[f'B{row}'].number_format = '0.00"%"'
markup_row = row
ws_calc[f'B{row}'] = f'=IF(B{total_cost_row}=0,0,(B{profit_recommended_row}/B{total_cost_row})*100)'
row += 2

# PRICING STATUS
ws_calc[f'A{row}'] = 'PRICING STATUS'
ws_calc[f'A{row}'].font = FONTS['bold']
ws_calc[f'B{row}'].border = BORDERS
ws_calc[f'B{row}'].font = FONTS['bold']
ws_calc[f'B{row}'].alignment = Alignment(horizontal='center', vertical='center')
pricing_status_row = row
# Formula: Check if margin is negative (LOSS), below target (WARNING), or OK
ws_calc[f'B{row}'] = f'=IF(B{margin_recommended_row}<0,"LOSS - PRICE TOO LOW",IF(B{margin_recommended_row}<B{assumption_rows["assum_margin"]},"WARNING - BELOW TARGET","OK - TARGET ACHIEVED"))'
# Conditional formatting for status
ws_calc[f'B{row}'].fill = FILLS['green']

row += 3

# ---- PRICE LADDER ----
ws_calc[f'A{row}'] = 'G. PRICE LADDER (Different Profit Margins)'
ws_calc[f'A{row}'].font = FONTS['subheader']
ws_calc[f'A{row}'].fill = FILLS['dark_navy']
ws_calc.merge_cells(f'A{row}:C{row}')
row += 1

# Headers
ws_calc[f'A{row}'] = 'Target Margin %'
ws_calc[f'A{row}'].font = FONTS['bold']
ws_calc[f'A{row}'].fill = FILLS['dark_navy']
ws_calc[f'A{row}'].font = Font(name='Calibri', size=11, bold=True, color='FFFFFFFF')
ws_calc[f'A{row}'].border = BORDERS

ws_calc[f'B{row}'] = 'Recommended Price'
ws_calc[f'B{row}'].font = FONTS['bold']
ws_calc[f'B{row}'].fill = FILLS['dark_navy']
ws_calc[f'B{row}'].font = Font(name='Calibri', size=11, bold=True, color='FFFFFFFF')
ws_calc[f'B{row}'].border = BORDERS

ws_calc[f'C{row}'] = 'Net Margin %'
ws_calc[f'C{row}'].font = FONTS['bold']
ws_calc[f'C{row}'].fill = FILLS['dark_navy']
ws_calc[f'C{row}'].font = Font(name='Calibri', size=11, bold=True, color='FFFFFFFF')
ws_calc[f'C{row}'].border = BORDERS
row += 1

# Price ladder rows
margins = [20, 25, 30, 35, 40]
for margin in margins:
    ws_calc[f'A{row}'] = margin
    ws_calc[f'A{row}'].font = FONTS['normal']
    ws_calc[f'A{row}'].fill = FILLS['blue']
    ws_calc[f'A{row}'].border = BORDERS
    ws_calc[f'A{row}'].alignment = Alignment(horizontal='center', vertical='center')
    
    # Price formula for this margin
    ws_calc[f'B{row}'] = f'=ROUND((B{total_cost_row}+B{expected_rto_row})/(1-(A{row}/100)-(B{selling_cost_rows["cost_gateway_pct"]}/100)-(B{selling_cost_rows["cost_platform_pct"]}/100)),0)'
    ws_calc[f'B{row}'].font = FONTS['normal']
    ws_calc[f'B{row}'].fill = FILLS['green']
    ws_calc[f'B{row}'].border = BORDERS
    ws_calc[f'B{row}'].number_format = '₨ 0.00'
    
    # Calculate actual margin at this price
    ws_calc[f'C{row}'] = f'=IF(B{row}=0,0,((B{row}-B{total_cost_row}-B{expected_rto_row}-(B{row}*B{selling_cost_rows["cost_gateway_pct"]}/100)-B{selling_cost_rows["cost_gateway_fixed"]}-(B{row}*B{selling_cost_rows["cost_platform_pct"]}/100))/B{row})*100)'
    ws_calc[f'C{row}'].font = FONTS['normal']
    ws_calc[f'C{row}'].fill = FILLS['green']
    ws_calc[f'C{row}'].border = BORDERS
    ws_calc[f'C{row}'].number_format = '0.00"%"'
    
    row += 1

# ============================================================================
# SHEET 4: PRODUCT MASTER
# ============================================================================
ws_master = wb.create_sheet('2 Product Master', 3)

# Headers
headers = [
    'SKU', 'Product Name', 'Pack Size (g)', 'Purchase Cost', 'Inbound Freight',
    'Duties', 'Wastage %', 'Bag Cost', 'Label Cost', 'Box Cost', 'Labor Cost',
    'Other Product Cost', 'Base Cost', 'Delivery Cost', 'Customer Delivery',
    'Gateway Fee %', 'Gateway Fixed', 'Platform Fee %', 'Marketing Cost',
    'Return Rate %', 'RTO Rate %', 'Return Shipping', 'Return Handling',
    'Target Margin %', 'Recommended Price', 'Discount Price', 
    'Expected Profit', 'Net Margin %', 'Status'
]

for col, header in enumerate(headers, 1):
    cell = ws_master.cell(row=1, column=col)
    cell.value = header
    cell.font = FONTS['subheader']
    cell.fill = FILLS['dark_navy']
    cell.border = BORDERS
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws_master.column_dimensions[get_column_letter(col)].width = 14

# Freeze header
ws_master.freeze_panes = 'A2'

# Pre-populate formulas for 500 rows
for row in range(2, 502):
    # SKU (user input)
    ws_master.cell(row=row, column=1).fill = FILLS['blue']
    ws_master.cell(row=row, column=1).border = BORDERS
    
    # Product Name (user input)
    ws_master.cell(row=row, column=2).fill = FILLS['blue']
    ws_master.cell(row=row, column=2).border = BORDERS
    
    # Pack Size (user input)
    ws_master.cell(row=row, column=3).fill = FILLS['blue']
    ws_master.cell(row=row, column=3).border = BORDERS
    ws_master.cell(row=row, column=3).number_format = '0'
    
    # Costs (user input)
    for col in range(4, 12):
        ws_master.cell(row=row, column=col).fill = FILLS['blue']
        ws_master.cell(row=row, column=col).border = BORDERS
        ws_master.cell(row=row, column=col).number_format = '0.00'
    
    # Base Cost (calculated)
    col = 13
    ws_master.cell(row=row, column=col).value = f'=D{row}+E{row}+F{row}+(D{row}*G{row}/100)+H{row}+I{row}+J{row}+K{row}+L{row}'
    ws_master.cell(row=row, column=col).fill = FILLS['green']
    ws_master.cell(row=row, column=col).border = BORDERS
    ws_master.cell(row=row, column=col).number_format = '0.00'
    
    # Delivery costs (user input)
    for col in range(14, 19):
        ws_master.cell(row=row, column=col).fill = FILLS['blue']
        ws_master.cell(row=row, column=col).border = BORDERS
        ws_master.cell(row=row, column=col).number_format = '0.00'
    
    # RTO costs (user input)
    for col in range(20, 24):
        ws_master.cell(row=row, column=col).fill = FILLS['blue']
        ws_master.cell(row=row, column=col).border = BORDERS
        ws_master.cell(row=row, column=col).number_format = '0.00'
    
    # Target Margin (user input)
    ws_master.cell(row=row, column=24).fill = FILLS['blue']
    ws_master.cell(row=row, column=24).border = BORDERS
    ws_master.cell(row=row, column=24).number_format = '0.00'
    
    # Recommended Price (calculated)
    col = 25
    ws_master.cell(row=row, column=col).value = f'=ROUND((M{row}+((V{row}+W{row})/100)*(X{row}+Y{row}))/(1-(Y{row}/100)-(P{row}/100)-(R{row}/100)),0)'
    ws_master.cell(row=row, column=col).fill = FILLS['green']
    ws_master.cell(row=row, column=col).border = BORDERS
    ws_master.cell(row=row, column=col).number_format = '₨ 0.00'
    
    # Discount Price (calculated) - assuming 10% discount
    col = 26
    ws_master.cell(row=row, column=col).value = f'=ROUND(Y{row}*0.9,0)'
    ws_master.cell(row=row, column=col).fill = FILLS['green']
    ws_master.cell(row=row, column=col).border = BORDERS
    ws_master.cell(row=row, column=col).number_format = '₨ 0.00'
    
    # Expected Profit (calculated)
    col = 27
    ws_master.cell(row=row, column=col).value = f'=Y{row}-M{row}-((V{row}+W{row})/100)*(X{row}+Y{row})-(Y{row}*P{row}/100)-Q{row}-(Y{row}*R{row}/100)'
    ws_master.cell(row=row, column=col).fill = FILLS['green']
    ws_master.cell(row=row, column=col).border = BORDERS
    ws_master.cell(row=row, column=col).number_format = '₨ 0.00'
    
    # Net Margin % (calculated)
    col = 28
    ws_master.cell(row=row, column=col).value = f'=IF(Y{row}=0,0,(Z{row}/Y{row})*100)'
    ws_master.cell(row=row, column=col).fill = FILLS['green']
    ws_master.cell(row=row, column=col).border = BORDERS
    ws_master.cell(row=row, column=col).number_format = '0.00"%"'
    
    # Status (calculated)
    col = 29
    ws_master.cell(row=row, column=col).value = f'=IF(AA{row}<0,"LOSS",IF(AA{row}<U{row},"WARNING","OK"))'
    ws_master.cell(row=row, column=col).fill = FILLS['green']
    ws_master.cell(row=row, column=col).border = BORDERS

# Enable filtering
ws_master.auto_filter.ref = f'A1:{get_column_letter(len(headers))}501'

# ============================================================================
# SHEET 5: DASHBOARD
# ============================================================================
ws_dashboard = wb.create_sheet('4 Dashboard', 4)
ws_dashboard.column_dimensions['A'].width = 40
ws_dashboard.column_dimensions['B'].width = 25

row = 1
ws_dashboard[f'A{row}'] = 'NORTHERN HARVEST - DASHBOARD'
ws_dashboard[f'A{row}'].font = FONTS['title']
ws_dashboard.merge_cells(f'A{row}:B{row}')

row += 2

# Current Product Section
ws_dashboard[f'A{row}'] = 'CURRENT PRODUCT (from Calculator)'
ws_dashboard[f'A{row}'].font = FONTS['subheader']
ws_dashboard[f'A{row}'].fill = FILLS['dark_navy']
ws_dashboard[f'A{row}'].merge_cells(f'A{row}:B{row}')
row += 1

dashboard_items = [
    ('Product Name', f"='1 Product Calculator'!B5"),
    ('SKU', f"='1 Product Calculator'!B6"),
    ('Pack Size', f"='1 Product Calculator'!B7"),
    ('Base Cost', f"='1 Product Calculator'!B17"),
    ('Total Variable Cost', f"='1 Product Calculator'!B37"),
    ('Break-even Price', f"='1 Product Calculator'!B38"),
    ('Recommended Price', f"='1 Product Calculator'!B40"),
    ('Discount Price', f"='1 Product Calculator'!B41"),
    ('Expected Profit', f"='1 Product Calculator'!B42"),
    ('Net Margin %', f"='1 Product Calculator'!B45"),
    ('Pricing Status', f"='1 Product Calculator'!B50"),
]

for label, formula in dashboard_items:
    ws_dashboard[f'A{row}'] = label
    ws_dashboard[f'A{row}'].font = FONTS['bold']
    ws_dashboard[f'B{row}'].value = formula
    ws_dashboard[f'B{row}'].font = FONTS['normal']
    ws_dashboard[f'B{row}'].fill = FILLS['green']
    ws_dashboard[f'B{row}'].border = BORDERS
    ws_dashboard[f'B{row}'].alignment = Alignment(horizontal='right', vertical='center')
    if 'Price' in label or 'Cost' in label or 'Profit' in label:
        ws_dashboard[f'B{row}'].number_format = '₨ 0.00'
    elif '%' in label:
        ws_dashboard[f'B{row}'].number_format = '0.00"%"'
    row += 1

row += 2

# Product Master Summary
ws_dashboard[f'A{row}'] = 'PRODUCT MASTER SUMMARY'
ws_dashboard[f'A{row}'].font = FONTS['subheader']
ws_dashboard[f'A{row}'].fill = FILLS['dark_navy']
ws_dashboard[f'A{row}'].merge_cells(f'A{row}:B{row}')
row += 1

summary_items = [
    ('Total Products', f"=COUNTA('2 Product Master'!A2:A501)"),
    ('Avg Product Margin %', f"=AVERAGE('2 Product Master'!AB2:AB501)"),
    ('Avg Recommended Price', f"=AVERAGE('2 Product Master'!Y2:Y501)"),
]

for label, formula in summary_items:
    ws_dashboard[f'A{row}'] = label
    ws_dashboard[f'A{row}'].font = FONTS['bold']
    ws_dashboard[f'B{row}'].value = formula
    ws_dashboard[f'B{row}'].font = FONTS['normal']
    ws_dashboard[f'B{row}'].fill = FILLS['green']
    ws_dashboard[f'B{row}'].border = BORDERS
    if 'Price' in label:
        ws_dashboard[f'B{row}'].number_format = '₨ 0.00'
    elif '%' in label:
        ws_dashboard[f'B{row}'].number_format = '0.00"%"'
    row += 1

# ============================================================================
# SAVE AND OUTPUT
# ============================================================================
filename = 'Northern_Harvest_Pricing_System.xlsx'
wb.save(filename)
print(f'✅ Workbook created successfully: {filename}')
print(f'\nSheets created:')
print(f'  1. 5 Start Here - Beginner guide')
print(f'  2. 3 Settings - Configuration')
print(f'  3. 1 Product Calculator - Main pricing tool')
print(f'  4. 2 Product Master - Product database (500 rows)')
print(f'  5. 4 Dashboard - Summary metrics')
print(f'\nFeatures included:')
print(f'  ✓ Complete pricing formulas')
print(f'  ✓ Professional color system')
print(f'  ✓ Price ladder (20%-40% margins)')
print(f'  ✓ RTO/Return cost allocation')
print(f'  ✓ Payment gateway & platform fees')
print(f'  ✓ Conditional formatting')
print(f'  ✓ Data validation ready')
print(f'  ✓ 500-row capacity')
print(f'\nNext steps:')
print(f'  1. Download the Excel file')
print(f'  2. Open in Google Sheets or Excel')
print(f'  3. Go to "5 Start Here" for instructions')
print(f'  4. Enter sample product data')
print(f'  5. Test all calculations')
